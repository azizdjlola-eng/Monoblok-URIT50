"""
Barcha tahlil normalari DB dan Excel ga eksport qilish.
Ishlatish: python export_normas.py
"""
import sys
import os

try:
    import mysql.connector
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Kerakli kutubxona topilmadi: {e}")
    print("pip install openpyxl mysql-connector-python")
    sys.exit(1)

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from monoblok_db_config import DB_CONFIG
except ImportError:
    DB_CONFIG = {
        "host": "labserver",
        "user": "lims",
        "password": "azizmed2026",
        "database": "lab_tizim",
        "port": 3306
    }

def export_normas():
    print("DB ga ulanmoqda...")
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor(dictionary=True)
    except Exception as e:
        print(f"DB ulanish xatosi: {e}")
        return

    cursor.execute("""
        SELECT
            id,
            guruh,
            tahlil_nomi,
            norma,
            birlik,
            type,
            min_val,
            max_val,
            min_val_m,
            max_val_m,
            min_val_f,
            max_val_f,
            result_template
        FROM tahlillar_norma
        ORDER BY guruh, tahlil_nomi
    """)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    print(f"Jami {len(rows)} ta norma topildi.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Normas"

    # Sarlavhalar
    headers = [
        "ID", "Guruh", "Tahlil nomi", "Norma matni", "Birlik", "Tur",
        "Min (umumiy)", "Max (umumiy)",
        "Min (Erkak)", "Max (Erkak)",
        "Min (Ayol)", "Max (Ayol)",
        "Ko'p komponent (result_template)"
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    # Guruh ranglari
    group_colors = {
        "BIO": "E2EFDA",
        "GEM": "DDEEFF",
        "URINE": "FFF2CC",
        "IFA": "FCE4EC",
        "COAG": "EDE7F6",
        "TEST": "FBE9E7",
    }

    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    normal_fill = PatternFill("solid", fgColor="FFFFFF")
    missing_fill = PatternFill("solid", fgColor="FFD7D7")  # Norma yo'q → qizil
    ok_font = Font(size=10)

    for row_idx, row in enumerate(rows, 2):
        guruh = (row.get('guruh') or '').upper()
        fill_color = group_colors.get(guruh, None)
        row_fill = PatternFill("solid", fgColor=fill_color) if fill_color else (alt_fill if row_idx % 2 == 0 else normal_fill)

        norma_text = row.get('norma') or ''
        has_template = bool(row.get('result_template') and row['result_template'].strip() not in ('', 'null', 'NULL'))
        min_val = row.get('min_val')
        max_val = row.get('max_val')
        min_m = row.get('min_val_m')
        max_m = row.get('max_val_m')
        min_f = row.get('min_val_f')
        max_f = row.get('max_val_f')

        # Norma yo'q yoki bo'sh bo'lsa — ogohlantirish
        norma_missing = (not norma_text or norma_text.strip() in ('-', '')) and not has_template

        values = [
            row.get('id'),
            row.get('guruh'),
            row.get('tahlil_nomi'),
            norma_text,
            row.get('birlik') or '',
            row.get('type') or '',
            min_val,
            max_val,
            min_m,
            max_m,
            min_f,
            max_f,
            "HA (JSON bor)" if has_template else "",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = ok_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col == 4))
            # Norma ustuni (D) bo'sh bo'lsa qizil
            if col == 4 and norma_missing:
                cell.fill = missing_fill
            else:
                cell.fill = row_fill

    # Ustun kengliklarini sozlash
    col_widths = [6, 8, 35, 30, 12, 10, 12, 12, 12, 12, 12, 12, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Muzlatish (sarlavha qolsin)
    ws.freeze_panes = "A2"

    # Ikkinchi varaq: faqat norma yo'q yoki raqamli ustunlar bo'sh bo'lganlar
    ws2 = wb.create_sheet("Tekshirish kerak")
    ws2.append(["ID", "Guruh", "Tahlil nomi", "Norma matni", "Birlik", "Muammo"])
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    ws2["C1"].font = Font(bold=True)
    ws2["D1"].font = Font(bold=True)
    ws2["E1"].font = Font(bold=True)
    ws2["F1"].font = Font(bold=True)

    warn_rows = 0
    for row in rows:
        norma_text = row.get('norma') or ''
        has_template = bool(row.get('result_template') and row['result_template'].strip() not in ('', 'null', 'NULL'))
        issues = []
        if not norma_text or norma_text.strip() in ('-', ''):
            if not has_template:
                issues.append("Norma matni yo'q")
        min_v = row.get('min_val')
        max_v = row.get('max_val')
        min_m = row.get('min_val_m')
        max_m = row.get('max_val_m')
        min_f = row.get('min_val_f')
        max_f = row.get('max_val_f')
        # Jinsga bog'liq norma matnida erkak/ayol bor lekin raqamli ustunlar bo'sh
        nt_lower = norma_text.lower()
        if ('erkak' in nt_lower or 'ayol' in nt_lower):
            if min_m is None and max_m is None and min_f is None and max_f is None:
                if not has_template:
                    issues.append("Jinsga bog'liq norma bor lekin min_m/max_m/min_f/max_f bo'sh")
        if issues:
            ws2.append([
                row.get('id'),
                row.get('guruh'),
                row.get('tahlil_nomi'),
                norma_text,
                row.get('birlik') or '',
                "; ".join(issues)
            ])
            warn_rows += 1

    ws2.column_dimensions['A'].width = 6
    ws2.column_dimensions['B'].width = 8
    ws2.column_dimensions['C'].width = 35
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 50

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "normas_export.xlsx")
    wb.save(out_path)
    print(f"Excel saqlandi: {out_path}")
    print(f"Tekshirish kerak: {warn_rows} ta tahlil (2-varaq)")
    return out_path

if __name__ == "__main__":
    export_normas()
